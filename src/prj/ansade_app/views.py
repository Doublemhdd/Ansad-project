from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction, models
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, timedelta
from datetime import timedelta
from django.db.models import Avg, Count


from openpyxl import Workbook, load_workbook

# On supprime l'import csv pour l'importation
# (on le garde si tu veux conserver l'export CSV)
import csv

from .models import (
    Wilaya,
    Moughataa,
    Commune,
    Product,
    PointOfSale,
    ProductPrice,
    Basket,
    BasketProduct,
    ProductType
)
from .forms import (
    WilayaForm,
    WilayaFilterForm,
    MoughataaForm,
    CommuneForm,
    ProductForm,
    PointOfSaleForm,
    ProductPriceForm,
    ProductPriceFilterForm,
    BasketForm,
    BasketProductForm,
    ProductTypeForm
)


# --------------------------------------------------
# PAGE D'ACCUEIL
# --------------------------------------------------
@login_required
def home(request):
    """
    Calculates the INPC for the last 4 months and prepares data for three charts:
    - A line chart showing INPC variation.
    - A pie chart showing product type distribution.
    - A bar chart showing the average price per product type (for the current date).
    """
    current_date = datetime.now()
    inpc_data = []      # List of dicts for each month: { 'month_label': "Month Year", 'inpc': value }
    labels_mois = []    # For the line chart
    valeurs_inpc = []   # For the line chart

    # Calculate the last 4 months (as (year, month) tuples)
    mois_a_calculer = []
    date_temp = current_date.replace(day=1)
    for i in range(4):
        mois_a_calculer.append((date_temp.year, date_temp.month))
        if date_temp.month == 1:
            date_temp = date_temp.replace(year=date_temp.year - 1, month=12)
        else:
            date_temp = date_temp.replace(month=date_temp.month - 1)
    mois_a_calculer.reverse()  # Oldest first

    ANNEE_BASE = 2019

    # Calculate global INPC for each month
    for annee, mois in mois_a_calculer:
        date_debut_mois = datetime(annee, mois, 1)
        if mois == 12:
            date_fin_mois = datetime(annee + 1, 1, 1) - timedelta(days=1)
        else:
            date_fin_mois = datetime(annee, mois + 1, 1) - timedelta(days=1)
        date_debut_base = datetime(ANNEE_BASE, mois, 1)
        if mois == 12:
            date_fin_base = datetime(ANNEE_BASE + 1, 1, 1) - timedelta(days=1)
        else:
            date_fin_base = datetime(ANNEE_BASE, mois + 1, 1) - timedelta(days=1)

        inpc_total = 0
        poids_global_total = 0

        for type_produit in ProductType.objects.all():
            produits = Product.objects.filter(product_type=type_produit)
            somme_indices = 0
            poids_groupe_total = 0

            for produit in produits:
                # Average base price for the product
                prix_base_obj = ProductPrice.objects.filter(
                    product=produit,
                    date_from__lte=date_fin_base,
                    date_to__gte=date_debut_base
                ).aggregate(Avg('price'))
                prix_base = prix_base_obj.get('price__avg')

                # Average current price for the product
                prix_courant_obj = ProductPrice.objects.filter(
                    product=produit,
                    date_from__lte=date_fin_mois,
                    date_to__gte=date_debut_mois
                ).aggregate(Avg('price'))
                prix_courant = prix_courant_obj.get('price__avg')

                # Average weight from BasketProduct for the product in the base year
                poids_obj = BasketProduct.objects.filter(
                    product=produit,
                    date_from__year=ANNEE_BASE,
                    date_to__year=ANNEE_BASE
                ).aggregate(Avg('weight'))
                poids = poids_obj.get('weight__avg') or 0

                if prix_base and prix_courant and poids > 0:
                    indice_produit = (prix_courant / prix_base) * 100
                    somme_indices += indice_produit * poids
                    poids_groupe_total += poids

            if poids_groupe_total > 0:
                indice_groupe = somme_indices / poids_groupe_total
                inpc_total += indice_groupe * (poids_groupe_total / 100)
                poids_global_total += poids_groupe_total

        global_inpc = inpc_total * (100 / poids_global_total) if poids_global_total > 0 else 0
        month_label = date_debut_mois.strftime('%B %Y')
        inpc_data.append({
            'year': annee,
            'month': mois,
            'month_label': month_label,
            'inpc': round(global_inpc, 2)
        })
        labels_mois.append(month_label)
        valeurs_inpc.append(round(global_inpc, 2))

    # Prepare data for a pie chart: distribution of product types by product count
    product_types_data = ProductType.objects.annotate(
        product_count=Count('product')
    ).values('name', 'product_count')
    pie_labels = [item['name'] for item in product_types_data]
    pie_data = [item['product_count'] for item in product_types_data]

    # Prepare data for a bar chart: average price per product type for the current date
    bar_labels = []
    bar_data = []
    for pt in ProductType.objects.all():
        avg_price_obj = ProductPrice.objects.filter(
            product__product_type=pt,
            date_from__lte=current_date,
            date_to__gte=current_date
        ).aggregate(Avg('price'))
        avg_price = avg_price_obj.get('price__avg') or 0
        bar_labels.append(pt.name)
        bar_data.append(round(avg_price, 2))

    context = {
        'inpc_data': inpc_data,
        'labels_mois': labels_mois,
        'valeurs_inpc': valeurs_inpc,
        'pie_labels': pie_labels,
        'pie_data': pie_data,
        'bar_labels': bar_labels,
        'bar_data': bar_data,
    }
    return render(request, 'home.html', context)



# --------------------------------------------------
# GESTION DES WILAYAS
# --------------------------------------------------
@login_required
def wilaya_list(request):
    wilayas = Wilaya.objects.all()
    filter_form = WilayaFilterForm(request.GET)

    if filter_form.is_valid():
        code = filter_form.cleaned_data.get('code')
        name = filter_form.cleaned_data.get('name')
        if code:
            wilayas = wilayas.filter(code__icontains=code)
        if name:
            wilayas = wilayas.filter(name__icontains=name)

    return render(request, 'wilaya_list.html', {
        'wilayas': wilayas,
        'filter_form': filter_form,
    })


@login_required
def wilaya_create(request):
    """
    Création d'une Wilaya manuellement (via formulaire)
    OU import depuis Excel (bouton "import_excel").
    Colonnes Excel attendues (en-tête en ligne 1) :
        A: Code
        B: Nom
    """
    if request.method == 'POST':
        if 'import_excel' in request.POST and 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active  # ou nom de la feuille si besoin

            created_count = 0
            skipped_count = 0

            with transaction.atomic():
                # On suppose que la ligne 1 est l'entête
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # row = (Code, Nom) par exemple
                    code = row[0]
                    name = row[1]

                    # Vérification basique
                    if not code or not name:
                        skipped_count += 1
                        continue

                    # get_or_create pour éviter doublons selon code
                    wilaya, created = Wilaya.objects.get_or_create(
                        code=code,
                        defaults={'name': name}
                    )
                    if not created:
                        # Si elle existait déjà, on peut choisir de la mettre à jour
                        # wilaya.name = name
                        # wilaya.save()
                        pass
                    created_count += 1

            messages.success(
                request,
                f"Import Excel terminé ! {created_count} Wilaya(s) créées/mises à jour, {skipped_count} ignorée(s)."
            )
            return redirect('wilaya_list')

        elif 'save_wilaya' in request.POST:
            form = WilayaForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Wilaya créée avec succès.")
                return redirect('wilaya_list')
    else:
        form = WilayaForm()

    return render(request, 'wilaya_form.html', {'form': form})


@login_required
def wilaya_update(request, pk):
    wilaya = get_object_or_404(Wilaya, pk=pk)
    if request.method == 'POST':
        form = WilayaForm(request.POST, instance=wilaya)
        if form.is_valid():
            form.save()
            messages.success(request, "Wilaya mise à jour avec succès.")
            return redirect('wilaya_list')
    else:
        form = WilayaForm(instance=wilaya)
    return render(request, 'wilaya_form.html', {'form': form})


@login_required
def wilaya_delete(request, pk):
    wilaya = get_object_or_404(Wilaya, pk=pk)
    if request.method == 'POST':
        wilaya.delete()
        messages.success(request, "Wilaya supprimée avec succès.")
        return redirect('wilaya_list')
    return render(request, 'wilaya_confirm_delete.html', {'wilaya': wilaya})


# --------------------------------------------------
# GESTION DES MOUGHATAAS
# --------------------------------------------------
@login_required
def moughataa_list(request):
    moughataas = Moughataa.objects.all()
    return render(request, 'moughataa_list.html', {'moughataas': moughataas})


@login_required
def moughataa_create(request):
    """
    Création d'une Moughataa manuellement OU import depuis Excel.
    Colonnes Excel attendues (ligne 1 = en-tête) :
        A: Code
        B: Nom
        C: Wilaya (nom)
    """
    if request.method == 'POST':
        if 'import_excel' in request.POST and 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']

            # Vérifier l'extension du fichier
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "Format de fichier invalide. Veuillez importer un fichier Excel.")
                return redirect('moughataa_list')

            try:
                wb = load_workbook(excel_file)
                ws = wb.active
            except Exception:
                messages.error(request, "Erreur lors de la lecture du fichier Excel.")
                return redirect('moughataa_list')

            created_count = 0
            skipped_count = 0

            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) < 3:
                        skipped_count += 1
                        continue

                    code = str(row[0]).strip() if row[0] else None
                    name = str(row[1]).strip() if row[1] else None
                    wilaya_name = str(row[2]).strip().capitalize() if row[2] else None

                    if not code or not name or not wilaya_name:
                        skipped_count += 1
                        continue

                    # Chercher ou créer la Wilaya
                    wilaya, created = Wilaya.objects.get_or_create(name=wilaya_name)

                    # Vérifier si la Moughataa existe déjà
                    if Moughataa.objects.filter(code=code).exists():
                        skipped_count += 1
                        continue

                    # Création de la Moughataa
                    Moughataa.objects.create(code=code, name=name, wilaya=wilaya)
                    created_count += 1

            messages.success(
                request,
                f"Import Moughataas : {created_count} créées, {skipped_count} ignorées."
            )
            return redirect('moughataa_list')

        elif 'save_moughataa' in request.POST:
            form = MoughataaForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Moughataa créée avec succès.")
                return redirect('moughataa_list')
            else:
                messages.error(request, "Erreur dans le formulaire. Veuillez corriger les erreurs.")

    else:
        form = MoughataaForm()

    return render(request, 'moughataa_form.html', {'form': form})

@login_required
def moughataa_update(request, pk):
    moughataa = get_object_or_404(Moughataa, pk=pk)
    if request.method == 'POST':
        form = MoughataaForm(request.POST, instance=moughataa)
        if form.is_valid():
            form.save()
            messages.success(request, "Moughataa mise à jour avec succès.")
            return redirect('moughataa_list')
    else:
        form = MoughataaForm(instance=moughataa)
    return render(request, 'moughataa_form.html', {'form': form})


@login_required
def moughataa_delete(request, pk):
    moughataa = get_object_or_404(Moughataa, pk=pk)
    if request.method == 'POST':
        moughataa.delete()
        messages.success(request, "Moughataa supprimée avec succès.")
        return redirect('moughataa_list')
    return render(request, 'moughataa_confirm_delete.html', {'moughataa': moughataa})


# --------------------------------------------------
# GESTION DES COMMUNES
# --------------------------------------------------
@login_required
def commune_list(request):
    communes = Commune.objects.all()
    return render(request, 'commune_list.html', {'communes': communes})



@login_required
def commune_create(request):
    """
    Vue pour créer une Commune via un formulaire ou un import Excel.
    """
    form = CommuneForm(request.POST or None)

    if request.method == 'POST':
        # 1) Cas de l'import Excel
        if 'import_excel' in request.POST and 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            
            # Lecture du fichier Excel avec openpyxl
            wb = load_workbook(excel_file)
            ws = wb.active  # ou wb["NomFeuille"] si besoin

            created_count = 0
            skipped_count = 0

            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Extraction des données (adaptation selon le fichier Excel)
                    code = str(row[0]).strip() if row[0] else None
                    name = str(row[1]).strip() if row[1] else None
                    moughataa_name = str(row[2]).strip().capitalize() if row[2] else None

                    if not code or not name or not moughataa_name:
                        skipped_count += 1
                        continue

                    # Vérifier si la Moughataa existe
                    try:
                        moughataa = Moughataa.objects.get(name=moughataa_name)
                    except Moughataa.DoesNotExist:
                        skipped_count += 1
                        continue

                    # Vérifier si la Commune existe déjà
                    if Commune.objects.filter(code=code).exists():
                        skipped_count += 1
                        continue

                    # Création de la Commune
                    Commune.objects.create(code=code, name=name, moughataa=moughataa)
                    created_count += 1

            messages.success(
                request,
                f"Import Excel terminé : {created_count} communes créées, {skipped_count} ignorées."
            )
            return redirect('commune_list')

        # 2) Cas du formulaire classique
        if form.is_valid():
            form.save()
            messages.success(request, "Commune créée avec succès.")
            return redirect('commune_list')

    return render(request, 'commune_form.html', {'form': form})



@login_required
def commune_update(request, pk):
    commune = get_object_or_404(Commune, pk=pk)
    if request.method == 'POST':
        form = CommuneForm(request.POST, instance=commune)
        if form.is_valid():
            form.save()
            messages.success(request, "Commune mise à jour avec succès.")
            return redirect('commune_list')
    else:
        form = CommuneForm(instance=commune)
    return render(request, 'commune_form.html', {'form': form})


@login_required
def commune_delete(request, pk):
    commune = get_object_or_404(Commune, pk=pk)
    if request.method == 'POST':
        commune.delete()
        messages.success(request, "Commune supprimée avec succès.")
        return redirect('commune_list')
    return render(request, 'commune_confirm_delete.html', {'commune': commune})

# --------------------------------------------------
# GESTION DES PRODUITS
# --------------------------------------------------
@login_required
def product_list(request):
    products = Product.objects.all()
    return render(request, 'product_list.html', {'products': products})


@login_required
def product_create(request):
    """
    Création d'un produit via formulaire
    OU import depuis un fichier Excel.
    """
    form = None  # Initialisation pour éviter UnboundLocalError
    if request.method == 'POST':
        # 1) Cas de l'import Excel
        if 'import_excel' in request.POST and 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active  # Ou wb["NomDeLaFeuille"] si nécessaire

            created_count = 0
            skipped_count = 0

            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Adapte les indices des colonnes selon ton fichier Excel
                    code = row[0] if len(row) > 0 else None
                    name = row[1] if len(row) > 1 else None
                    description = row[2] if len(row) > 2 else ''
                    unit_measure = row[3] if len(row) > 3 else ''
                    product_type_name = row[4] if len(row) > 4 else None

                    if not code or not name:
                        skipped_count += 1
                        continue

                    # Gestion de la FK "ProductType"
                    product_type = None
                    if product_type_name:
                        try:
                            product_type = ProductType.objects.get(name=product_type_name)
                        except ProductType.DoesNotExist:
                            skipped_count += 1
                            continue

                    # Création du produit
                    Product.objects.create(
                        code=code,
                        name=name,
                        description=description,
                        unit_measure=unit_measure,
                        product_type=product_type
                    )
                    created_count += 1

            messages.success(
                request,
                f"Import Excel terminé : {created_count} produits créés, {skipped_count} ignorés."
            )
            return redirect('product_list')

        # 2) Cas du formulaire classique
        elif 'save_product' in request.POST:
            form = ProductForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Produit créé avec succès.")
                return redirect('product_list')

    else:
        form = ProductForm()

    return render(request, 'product_form.html', {'form': form})

@login_required
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, "Produit mis à jour avec succès.")
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'product_form.html', {'form': form})


@login_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        messages.success(request, "Produit supprimé avec succès.")
        return redirect('product_list')
    return render(request, 'product_confirm_delete.html', {'product': product})


# --------------------------------------------------
# GESTION DES POINTS DE VENTE
# --------------------------------------------------
@login_required
def point_of_sale_list(request):
    points_of_sale = PointOfSale.objects.all()
    return render(request, 'point_of_sale_list.html', {'points_of_sale': points_of_sale})



@login_required
def point_of_sale_create(request):
    """
    Create a PointOfSale either via a standard form submission or by importing an Excel file.

    Expected Excel file structure (with header row, which is skipped):
      - Column A: Code (required)
      - Column B: Name (required)
      - Column C: Type (optional)
      - Column D: Commune code (required; used to look up the related Commune)
      - Column E: GPS Latitude (optional)
      - Column F: GPS Longitude (optional)
    """
    form = None  # Initialize to avoid UnboundLocalError

    if request.method == 'POST':
        # Case 1: Excel import
        if 'import_excel' in request.POST and 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            try:
                wb = load_workbook(excel_file)
            except Exception as e:
                messages.error(request, f"Error reading Excel file: {e}")
                return redirect('point_of_sale_list')
            ws = wb.active  # You can change this if a specific sheet is needed

            created_count = 0
            skipped_count = 0

            with transaction.atomic():
                # Skip header row (start at row 2)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Debug: print the raw row data
                    print("Row raw data:", row)
                    
                    # Extract and clean values:
                    code = str(row[0]).strip() if row[0] is not None else ""
                    name = str(row[1]).strip() if row[1] is not None else ""
                    pos_type = str(row[2]).strip() if row[2] is not None else ""
                    commune_code = str(row[3]).strip() if row[3] is not None else ""
                    
                    # Convert GPS latitude and longitude to float, if available
                    try:
                        gps_lat = float(row[4]) if row[4] not in (None, "") else None
                    except (ValueError, TypeError):
                        gps_lat = None
                    try:
                        gps_lon = float(row[5]) if row[5] not in (None, "") else None
                    except (ValueError, TypeError):
                        gps_lon = None

                    # Check required fields: code, name, and commune_code must be provided
                    if not code or not name or not commune_code:
                        skipped_count += 1
                        continue

                    # Lookup the Commune using the provided commune code
                    try:
                        commune_instance = Commune.objects.get(code=commune_code)
                    except Commune.DoesNotExist:
                        skipped_count += 1
                        continue

                    # Optionally, if a PointOfSale with the same code already exists, skip this row
                    if PointOfSale.objects.filter(code=code).exists():
                        skipped_count += 1
                        continue

                    # Create the PointOfSale instance
                    PointOfSale.objects.create(
                        code=code,
                        name=name,
                        type=pos_type,          # 'type' is a model field; we use pos_type to avoid conflict with built-in type
                        commune=commune_instance,
                        gps_lat=gps_lat,
                        gps_lon=gps_lon,
                    )
                    created_count += 1

            messages.success(
                request,
                f"Excel import completed: {created_count} point(s) de vente created, {skipped_count} skipped."
            )
            return redirect('point_of_sale_list')

        # Case 2: Standard form submission
        elif 'save_point_of_sale' in request.POST:
            form = PointOfSaleForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Point de vente créé avec succès.")
                return redirect('point_of_sale_list')
    else:
        form = PointOfSaleForm()

    return render(request, 'point_of_sale_form.html', {'form': form})


@login_required
def point_of_sale_update(request, pk):
    point_of_sale = get_object_or_404(PointOfSale, pk=pk)
    if request.method == 'POST':
        form = PointOfSaleForm(request.POST, instance=point_of_sale)
        if form.is_valid():
            form.save()
            messages.success(request, "Point de vente mis à jour avec succès.")
            return redirect('point_of_sale_list')
    else:
        form = PointOfSaleForm(instance=point_of_sale)
    return render(request, 'point_of_sale_form.html', {'form': form})


@login_required
def point_of_sale_delete(request, pk):
    point_of_sale = get_object_or_404(PointOfSale, pk=pk)
    if request.method == 'POST':
        point_of_sale.delete()
        messages.success(request, "Point de vente supprimé avec succès.")
        return redirect('point_of_sale_list')
    return render(request, 'point_of_sale_confirm_delete.html', {'point_of_sale': point_of_sale})


# --------------------------------------------------
# GESTION DES PRIX DES PRODUITS
# --------------------------------------------------
@login_required
def price_list(request):
    prices = ProductPrice.objects.all()
    filter_form = ProductPriceFilterForm(request.GET)

    if filter_form.is_valid():
        product = filter_form.cleaned_data.get('product')
        point_of_sale = filter_form.cleaned_data.get('point_of_sale')
        price_min = filter_form.cleaned_data.get('price_min')
        price_max = filter_form.cleaned_data.get('price_max')
        date_from = filter_form.cleaned_data.get('date_from')
        date_to = filter_form.cleaned_data.get('date_to')

        if product:
            prices = prices.filter(product__name__icontains=product)
        if point_of_sale:
            prices = prices.filter(point_of_sale__name__icontains=point_of_sale)
        if price_min is not None:
            prices = prices.filter(price__gte=price_min)
        if price_max is not None:
            prices = prices.filter(price__lte=price_max)
        if date_from:
            prices = prices.filter(date_from__gte=date_from)
        if date_to:
            prices = prices.filter(date_to__lte=date_to)

    return render(request, 'price_list.html', {
        'prices': prices,
        'filter_form': filter_form,
    })


@login_required
def price_create(request):
    """
    Création d'un prix via formulaire
    OU import depuis Excel (colonnes attendues, par ex.):
      - Colonne A : Produit (nom)
      - Colonne B : Point de vente (nom)
      - Colonne C : Prix
      - Colonne D : Date de début (YYYY-MM-DD)
      - Colonne E : Date de fin (YYYY-MM-DD)
    """
    form = None  # Pour éviter UnboundLocalError

    if request.method == 'POST':
        # Cas import Excel
        if 'import_excel' in request.POST and 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            try:
                wb = load_workbook(excel_file)
            except Exception as e:
                messages.error(request, f"Erreur lors de la lecture du fichier Excel: {e}")
                return redirect('price_list')
            ws = wb.active  # ou spécifiez une feuille si nécessaire

            created_count = 0
            skipped_count = 0

            with transaction.atomic():
                # Parcourir les lignes en partant de la 2ème ligne (en-tête ignoré)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    product_name = row[0]
                    pos_name = row[1]
                    price_str = row[2]
                    date_from_str = row[3]
                    date_to_str = row[4]

                    # Vérifier que les champs obligatoires sont présents
                    if not product_name or not pos_name or not price_str:
                        skipped_count += 1
                        continue

                    # Récupération du produit
                    try:
                        product = Product.objects.get(name=product_name)
                    except Product.DoesNotExist:
                        skipped_count += 1
                        continue

                    # Récupération du point de vente
                    # Utiliser filter() au lieu de get() pour éviter MultipleObjectsReturned
                    ps_qs = PointOfSale.objects.filter(name=pos_name)
                    if not ps_qs.exists():
                        skipped_count += 1
                        continue
                    point_of_sale = ps_qs.first()  # Choisit le premier objet trouvé

                    # Conversion du prix en Decimal
                    try:
                        price_value = Decimal(str(price_str))
                    except (InvalidOperation, TypeError):
                        skipped_count += 1
                        continue

                    # Parsing des dates
                    date_format = "%Y-%m-%d"
                    date_from_parsed = None
                    date_to_parsed = None

                    if date_from_str:
                        try:
                            date_from_parsed = datetime.strptime(str(date_from_str), date_format).date()
                        except ValueError:
                            pass

                    if date_to_str:
                        try:
                            date_to_parsed = datetime.strptime(str(date_to_str), date_format).date()
                        except ValueError:
                            pass

                    # Si date_from n'est pas parsable, sauter la ligne
                    if not date_from_parsed:
                        skipped_count += 1
                        continue

                    # Création du ProductPrice
                    ProductPrice.objects.create(
                        product=product,
                        point_of_sale=point_of_sale,
                        price=price_value,
                        date_from=date_from_parsed,
                        date_to=date_to_parsed
                    )
                    created_count += 1

            messages.success(
                request,
                f"Import des prix terminé : {created_count} créés, {skipped_count} ignorés."
            )
            return redirect('price_list')

        # Cas du formulaire standard
        elif 'save_price' in request.POST:
            form = ProductPriceForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Prix créé avec succès.")
                return redirect('price_list')
    else:
        form = ProductPriceForm()

    return render(request, 'price_form.html', {'form': form})

@login_required
def price_update(request, pk):
    price = get_object_or_404(ProductPrice, pk=pk)
    if request.method == 'POST':
        form = ProductPriceForm(request.POST, instance=price)
        if form.is_valid():
            form.save()
            messages.success(request, "Prix mis à jour avec succès.")
            return redirect('price_list')
    else:
        form = ProductPriceForm(instance=price)
    return render(request, 'price_form.html', {'form': form})


@login_required
def price_delete(request, pk):
    price = get_object_or_404(ProductPrice, pk=pk)
    if request.method == 'POST':
        price.delete()
        messages.success(request, "Prix supprimé avec succès.")
        return redirect('price_list')
    return render(request, 'price_confirm_delete.html', {'price': price})


# --------------------------------------------------
# GESTION DES PANIERS
# --------------------------------------------------
@login_required
def basket_list(request):
    baskets = Basket.objects.all()
    return render(request, 'basket_list.html', {'baskets': baskets})


@login_required
def basket_create(request):
    """
    Création d'un panier manuellement OU import depuis un fichier Excel.
    Colonnes Excel attendues (ligne d'en-tête en 1) :
        A: Nom
        B: Description
    """
    form = None  # Initialisation pour éviter UnboundLocalError

    if request.method == 'POST':
        if 'import_excel' in request.POST and 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active

            created_count = 0
            skipped_count = 0

            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    name = row[0]
                    description = row[1]

                    if not name:
                        skipped_count += 1
                        continue

                    Basket.objects.create(name=name, description=description)
                    created_count += 1

            messages.success(request, f"Import Excel terminé : {created_count} paniers créés, {skipped_count} ignorés.")
            return redirect('basket_list')

        elif 'save_basket' in request.POST:
            form = BasketForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Panier créé avec succès.")
                return redirect('basket_list')

    if form is None:  # S'assurer que `form` est toujours défini
        form = BasketForm()

    return render(request, 'basket_form.html', {'form': form})


@login_required
def basket_update(request, pk):
    basket = get_object_or_404(Basket, pk=pk)
    if request.method == 'POST':
        form = BasketForm(request.POST, instance=basket)
        if form.is_valid():
            form.save()
            messages.success(request, "Panier mis à jour avec succès.")
            return redirect('basket_list')
    else:
        form = BasketForm(instance=basket)
    return render(request, 'basket_form.html', {'form': form})


@login_required
def basket_delete(request, pk):
    basket = get_object_or_404(Basket, pk=pk)
    if request.method == 'POST':
        basket.delete()
        messages.success(request, "Panier supprimé avec succès.")
        return redirect('basket_list')
    return render(request, 'basket_confirm_delete.html', {'basket': basket})


# --------------------------------------------------
# GESTION DES PRODUITS DANS LES PANIERS
# --------------------------------------------------
@login_required
def basket_product_list(request):
    basket_products = BasketProduct.objects.all()
    return render(request, 'basket_product_list.html', {'basket_products': basket_products})


@login_required
def basket_product_create(request):
    """
    Création d'un BasketProduct via formulaire OU import depuis un fichier Excel.
    
    Expected Excel file structure (with header row, which is skipped):
      - Column A: Basket name (required)
      - Column B: Product name (required)
      - Column C: Weight (required, numeric)
      - Column D: Date de début (date_from) in format YYYY-MM-DD or as an Excel date (required)
      - Column E: Date de fin (date_to) in format YYYY-MM-DD or as an Excel date (optional)
    """
    form = None  # To avoid UnboundLocalError

    if request.method == 'POST':
        # Case 1: Excel Import
        if 'import_excel' in request.POST and 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            try:
                wb = load_workbook(excel_file)
            except Exception as e:
                messages.error(request, f"Erreur lors de la lecture du fichier Excel: {e}")
                return redirect('basket_product_list')
            ws = wb.active  # or wb["SheetName"] if you need a specific sheet

            created_count = 0
            skipped_count = 0
            date_format = "%Y-%m-%d"  # Expected date format for string values

            with transaction.atomic():
                # Iterate over the rows starting from row 2 (to skip the header)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Extract values and trim text fields
                    basket_name = str(row[0]).strip() if row[0] is not None else ""
                    product_name = str(row[1]).strip() if row[1] is not None else ""
                    weight_val = row[2]  # Expected to be numeric
                    
                    # Parse date_from (Column D)
                    if len(row) > 3 and row[3]:
                        if isinstance(row[3], datetime):
                            date_from_parsed = row[3].date()
                        else:
                            try:
                                date_from_parsed = datetime.strptime(str(row[3]).strip(), date_format).date()
                            except ValueError:
                                date_from_parsed = None
                    else:
                        date_from_parsed = None

                    # Parse date_to (Column E)
                    if len(row) > 4 and row[4]:
                        if isinstance(row[4], datetime):
                            date_to_parsed = row[4].date()
                        else:
                            try:
                                date_to_parsed = datetime.strptime(str(row[4]).strip(), date_format).date()
                            except ValueError:
                                date_to_parsed = None
                    else:
                        date_to_parsed = None

                    # Validate that required fields are present
                    if not basket_name or not product_name or weight_val is None or not date_from_parsed:
                        skipped_count += 1
                        continue

                    # Retrieve the Basket instance by name
                    try:
                        basket = Basket.objects.get(name=basket_name)
                    except Basket.DoesNotExist:
                        skipped_count += 1
                        continue

                    # Retrieve the Product instance by name
                    try:
                        product = Product.objects.get(name=product_name)
                    except Product.DoesNotExist:
                        skipped_count += 1
                        continue

                    # Convert weight to float
                    try:
                        weight_converted = float(weight_val)
                    except (ValueError, TypeError):
                        skipped_count += 1
                        continue

                    # Create the BasketProduct instance with parsed dates from Excel
                    BasketProduct.objects.create(
                        basket=basket,
                        product=product,
                        weight=weight_converted,
                        date_from=date_from_parsed,
                        date_to=date_to_parsed
                    )
                    created_count += 1

            messages.success(
                request,
                f"Import Excel terminé : {created_count} lignes créées, {skipped_count} ignorées."
            )
            return redirect('basket_product_list')

        # Case 2: Standard form submission
        elif 'save_basket_product' in request.POST:
            form = BasketProductForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Produit ajouté au panier avec succès.")
                return redirect('basket_product_list')
    else:
        form = BasketProductForm()

    return render(request, 'basket_product_form.html', {'form': form})


@login_required
def basket_product_update(request, pk):
    basket_product = get_object_or_404(BasketProduct, pk=pk)
    if request.method == 'POST':
        form = BasketProductForm(request.POST, instance=basket_product)
        if form.is_valid():
            form.save()
            messages.success(request, "Produit du panier mis à jour avec succès.")
            return redirect('basket_product_list')
    else:
        form = BasketProductForm(instance=basket_product)
    return render(request, 'basket_product_form.html', {'form': form})


@login_required
def basket_product_delete(request, pk):
    basket_product = get_object_or_404(BasketProduct, pk=pk)
    if request.method == 'POST':
        basket_product.delete()
        messages.success(request, "Produit du panier supprimé avec succès.")
        return redirect('basket_product_list')
    return render(request, 'basket_product_confirm_delete.html', {'basket_product': basket_product})


# --------------------------------------------------
# GESTION DES TYPES DE PRODUITS
# --------------------------------------------------
@login_required
def product_type_list(request):
    product_types = ProductType.objects.all()
    return render(request, 'product_type_list.html', {'product_types': product_types})


@login_required
def product_type_create(request):
    """
    Création d'un type de produit via formulaire
    OU import depuis un fichier Excel.
    Structure Excel supposée :
      - Colonne A : Nom du type de produit
      - (Colonne B : Description, si nécessaire)
    """
    form = None  # Initialisation pour éviter UnboundLocalError
    if request.method == 'POST':
        # 1) Cas import Excel
        if 'import_excel' in request.POST and 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active  # ou wb["NomFeuille"] si besoin

            created_count = 0
            skipped_count = 0

            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # row[0] = name
                    # row[1] = description (si tu l'as)
                    name = row[0]
                    description = row[1] if len(row) > 1 else ''

                    # Vérification minimale
                    if not name:
                        skipped_count += 1
                        continue

                    ProductType.objects.create(
                        name=name,
                        description=description
                    )
                    created_count += 1

            messages.success(
                request,
                f"Import Excel terminé : {created_count} type(s) de produit créés, {skipped_count} ignorés."
            )
            return redirect('product_type_list')

        # 2) Cas formulaire standard
        elif 'save_product_type' in request.POST:
            form = ProductTypeForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Type de produit créé avec succès.")
                return redirect('product_type_list')
    else:
        form = ProductTypeForm()

    return render(request, 'product_type_form.html', {'form': form})

@login_required
def product_type_update(request, pk):
    product_type = get_object_or_404(ProductType, pk=pk)
    if request.method == 'POST':
        form = ProductTypeForm(request.POST, instance=product_type)
        if form.is_valid():
            form.save()
            messages.success(request, "Type de produit mis à jour avec succès.")
            return redirect('product_type_list')
    else:
        form = ProductTypeForm(instance=product_type)
    return render(request, 'product_type_form.html', {'form': form})


@login_required
def product_type_delete(request, pk):
    product_type = get_object_or_404(ProductType, pk=pk)
    if request.method == 'POST':
        product_type.delete()
        messages.success(request, "Type de produit supprimé avec succès.")
        return redirect('product_type_list')
    return render(request, 'product_type_confirm_delete.html', {'product_type': product_type})


# --------------------------------------------------
# EXPORT DES DONNÉES (CSV / Excel)
# --------------------------------------------------
@login_required
def export_wilaya_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="wilayas.csv"'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Code', 'Nom'])
    wilayas = Wilaya.objects.all()
    for wilaya in wilayas:
        writer.writerow([wilaya.id, wilaya.code, wilaya.name])
    return response


@login_required
def export_moughataa_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="moughataas.csv"'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Code', 'Nom', 'Wilaya'])
    moughataas = Moughataa.objects.select_related('wilaya').all()
    for moughataa in moughataas:
        writer.writerow([moughataa.id, moughataa.code, moughataa.name, moughataa.wilaya.name])
    return response


@login_required
def export_price_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="prices.csv"'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Produit', 'Point de vente', 'Prix', 'Date de début', 'Date de fin'])
    prices = ProductPrice.objects.select_related('product', 'point_of_sale').all()
    for price in prices:
        writer.writerow([
            price.id,
            price.product.name,
            price.point_of_sale.name,
            price.price,
            price.date_from,
            price.date_to
        ])
    return response


@login_required
def export_basket_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="baskets.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Paniers"
    ws.append(['ID', 'Nom', 'Description'])
    baskets = Basket.objects.all()
    for basket in baskets:
        ws.append([basket.id, basket.name, basket.description])
    wb.save(response)
    return response

@login_required
def export_commune_excel(request):
    response = HttpResponse(
        content_type='application/ms-excel'
    )
    response['Content-Disposition'] = 'attachment; filename="communes.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Communes"

    # Entêtes
    ws.append(['ID', 'Nom'])

    # Contenu
    for commune in Commune.objects.all():
        ws.append([commune.id, commune.name])

    wb.save(response)
    return response

@login_required
def export_product_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="products.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Produits"

    # Entêtes (sans 'Prix' qui n'existe pas)
    ws.append(['ID', 'Code', 'Nom', 'Description', 'Unité de mesure', 'Type de produit'])

    products = Product.objects.select_related('product_type').all()
    for product in products:
        ws.append([
            product.id,
            product.code,
            product.name,
            product.description or '',
            product.unit_measure or '',
            product.product_type.name if product.product_type else ''
        ])

    wb.save(response)
    return response

@login_required
def export_point_of_sale_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="points_of_sale.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Points de vente"

    # Ligne d'entête
    ws.append(['ID', 'Nom', 'Adresse', 'Ville'])

    points_of_sale = PointOfSale.objects.all()
    for pos in points_of_sale:
        ws.append([
            pos.id,
            pos.name,
            pos.address or '',
            pos.city or ''
        ])

    wb.save(response)
    return response

@login_required
def export_product_type_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="product_types.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Types de produits"

    # Entête
    ws.append(['ID', 'Nom', 'Description'])

    for pt in ProductType.objects.all():
        ws.append([
            pt.id,
            pt.name,
            pt.description or ''
        ])

    wb.save(response)
    return response

@login_required
def export_basket_product_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="basket_products.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Produits dans Paniers"

    # Entêtes
    ws.append(['ID', 'Basket', 'Product', 'Quantity'])

    basket_products = BasketProduct.objects.select_related('basket', 'product').all()
    for bp in basket_products:
        ws.append([
            bp.id,
            bp.basket.name if bp.basket else '',
            bp.product.name if bp.product else '',
            bp.quantity
        ])

    wb.save(response)
    return response

#calcule de l'INPC
@login_required
def calculer_inpc(request):
    """
    Calcule l'Indice National des Prix à la Consommation (INPC)
    avec possibilité de sélectionner l'année et le mois
    """
    # Année de base
    ANNEE_BASE = 2019
    
    # Récupérer l'année et le mois à partir de la requête, sinon utiliser l'année et le mois actuels
    annee_courante = request.GET.get('annee', datetime.now().year)
    mois_courant = request.GET.get('mois', datetime.now().month)
    
    try:
        annee_courante = int(annee_courante)
        mois_courant = int(mois_courant)
    except ValueError:
        # En cas de valeurs invalides, utiliser l'année et le mois actuels
        annee_courante = datetime.now().year
        mois_courant = datetime.now().month
    
    # Récupérer tous les types de produits (catégories COICOP)
    types_produits = ProductType.objects.all()
    
    # Dictionnaires pour stocker les résultats
    inpc_par_groupe = []
    inpc_global = {}
    
    # Calculer l'INPC pour chaque type de produit
    for type_produit in types_produits:
        # Récupérer les produits de ce type
        produits = Product.objects.filter(product_type=type_produit)
        
        # Variables pour stocker les totaux
        prix_total_base = 0
        prix_total_courant = 0
        poids_total = 0
        produits_calcules = 0
        
        for produit in produits:
            # Récupérer les prix pour l'année de base
            prix_base = ProductPrice.objects.filter(
                product=produit,
                date_from__year=ANNEE_BASE,
                date_from__month=mois_courant
            ).order_by('-date_from').first()
            
            # Récupérer les prix pour l'année courante
            prix_courant = ProductPrice.objects.filter(
                product=produit,
                date_from__year=annee_courante,
                date_from__month=mois_courant
            ).order_by('-date_from').first()
            
            # Récupérer le poids du produit dans les paniers
            cart_products = BasketProduct.objects.filter(
                   product=produit,
           date_from__year=annee_courante,
             date_to__year=annee_courante
            )

            
            # Vérifier que tous les éléments nécessaires sont présents
            if prix_base and prix_courant and cart_products:
                poids_moyen = cart_products.aggregate(
                    models.Avg('weighting')
                )['weighting__avg'] or 0
                
                # Ajouter au total uniquement si le poids est significatif
                if poids_moyen > 0:
                    prix_total_base += prix_base.value * poids_moyen
                    prix_total_courant += prix_courant.value * poids_moyen
                    poids_total += poids_moyen
                    produits_calcules += 1
        
        # Calculer l'INPC pour ce groupe de produits
        # INPC = (Prix courant / Prix base) * 100
        inpc_groupe = (prix_total_courant / prix_total_base * 100) if prix_total_base > 0 else 0
        
        # N'ajouter que les groupes avec des produits calculés
        if produits_calcules > 0:
            inpc_par_groupe.append({
                'Année': annee_courante,
                'Mois': mois_courant,
                'Groupe': type_produit.label,
                'INPC': inpc_groupe,
                'Produits Calculés': produits_calcules
            })
    
    # Calculer l'INPC global
    if inpc_par_groupe:
        inpc_total = sum(groupe['INPC'] for groupe in inpc_par_groupe) / len(inpc_par_groupe)
    else:
        inpc_total = 0
    
    inpc_global[(annee_courante, mois_courant)] = inpc_total
    
    # Préparer les options d'années et de mois pour le formulaire
    annees_disponibles = sorted(set(
        ProductPrice.objects.values_list('date_from__year', flat=True).distinct()
    ))
    
    context = {
        'annee_base': ANNEE_BASE,
        'annee_courante': annee_courante,
        'mois_courant': mois_courant,
        'annees_disponibles': annees_disponibles,
        'inpc_par_groupe': inpc_par_groupe,
        'inpc_global': inpc_global
    }
    
    return render(request, 'calculer_inpc.html', context)